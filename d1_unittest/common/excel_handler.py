# 1，打开Excel文件 （路径+文件名）
# 2，获取表单
# 3，使用行号和列号去确定需要读取的数据
# 4，关闭文件
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler:
    """操作Excel"""

    def __init__(self, file):
        self.file = file

    def open_sheet(self, name) -> Worksheet:
        """打开表单"""
        wb = load_workbook(self.file)
        sheet = wb[name]
        wb.close()
        return sheet

    def header(self, sheet_name):
        """获取表头"""
        sheet = self.open_sheet(sheet_name)
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers

    def read(self, sheet_name):
        """读取所有的数据"""
        sheet = self.open_sheet(sheet_name)
        rows = list(sheet.rows)[1:]
        data = []
        for row in rows:  # 循环取rows中的每行
            row_data = []  # 存储一行的数据
            for cell in row:  # 循环取 row的第一个单元格
                row_data.append(cell.value)  # 将每个单元格的数据返回row_data
                # 列表转成字典，需要和header 去 zip
                data_dict = dict(zip(self.header(sheet_name), row_data))
            data.append(data_dict)  # 将每行数据返回data
        return data  # 存储所有数据

    @staticmethod
    def write(file, sheet_name, row, column, data):
        """写入 Excel 数据"""
        wb = load_workbook(file)
        sheet = wb[sheet_name]
        # 修改单元格
        sheet.cell(row, column).value = data
        # 保存
        wb.save(file)
        # 关闭
        wb.close()


if __name__ == '__main__':
    excel = ExcelHandler(r'H:\a\cases.xlsx')
    excel.write(r'H:\a\cases.xlsx', 'Sheet1', 2, 1, '8888888')

    # data = excel.read('Sheet1')
    # print(data)
